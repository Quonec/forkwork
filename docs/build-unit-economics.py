# -*- coding: utf-8 -*-
"""ForkWork - Юнит-экономика v2 (вычисляемая снизу вверх).

Каждый показатель — формула от листа «Допущения»: меняешь драйвер → пересчёт.
Синие ячейки — ввод. Генерирует docs/ForkWork-Unit-Economics-v2.xlsx.
Запуск: python build-unit-economics.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OCHRE = "C9822F"; OCHRE_D = "9F651E"; CREAM = "F7F1E4"; INK = "241C10"
INPUT_FILL = PatternFill("solid", fgColor="DCE9F7"); INPUT_FONT = Font(color="1F4E79", bold=True)
HDR_FILL = PatternFill("solid", fgColor=OCHRE); HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
SEC_FILL = PatternFill("solid", fgColor=CREAM); SEC_FONT = Font(color=OCHRE_D, bold=True, size=11)
RES_FONT = Font(bold=True, size=12, color=OCHRE_D)
THIN = Side(style="thin", color="E5D8BD"); BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_RUB = '#,##0" ₽"'; F_RUB2 = '#,##0.0" ₽"'; F_PCT = '0.0%'; F_X = '0.0"×"'; F_MO = '0.0" мес"'; F_NUM = '#,##0'

REF = {}  # key -> "'Допущения'!$C$n"

def sheet_header(ws, title, sub):
    ws["A1"] = title; ws["A1"].font = Font(bold=True, size=14, color=OCHRE_D)
    ws["A2"] = sub; ws["A2"].font = Font(size=9, color="6D5E45")
    ws.sheet_view.showGridLines = False

def put(ws, r, label, value, key=None, fmt=F_NUM, is_input=False, note=""):
    ws.cell(r, 2, label).alignment = Alignment(wrap_text=True)
    c = ws.cell(r, 3, value); c.number_format = fmt; c.border = BORDER
    if is_input: c.fill = INPUT_FILL; c.font = INPUT_FONT
    if note: ws.cell(r, 4, note).font = Font(size=8, color="9B8A6C")
    if key: REF[key] = f"'Допущения'!$C${r}"
    return r + 1

def section(ws, r, text):
    ws.cell(r, 2, text).fill = SEC_FILL; ws.cell(r, 2).font = SEC_FONT
    ws.cell(r, 3).fill = SEC_FILL; ws.cell(r, 4).fill = SEC_FILL
    return r + 1

wb = Workbook()

# ============ Обложка ============
ws = wb.active; ws.title = "Обложка"; ws.sheet_view.showGridLines = False
ws["B3"] = "ForkWork"; ws["B3"].font = Font(bold=True, size=26, color=OCHRE_D)
ws["B4"] = "Юнит-экономика v2 — вычисляемая модель"; ws["B4"].font = Font(bold=True, size=14)
ws["B6"] = "Принцип: ни одной экспертной константы в результатах."
ws["B7"] = "LTV, CAC, take-rate и окупаемость района выводятся формулами из листа «Допущения»."
ws["B9"] = "Синяя ячейка = ввод (меняйте). Белая = формула (не трогать)."
ws["B9"].fill = INPUT_FILL; ws["B9"].font = INPUT_FONT
ws["B11"] = "Листы: Допущения → Выручка на заказ → CAC → Когорты LTV → Юнит-экономика → Экономика района → Чувствительность"
ws["B13"] = "Дисклеймер: модель иллюстративная, для планирования пилота; факты пилота заменят допущения."
ws["B13"].font = Font(italic=True, size=9, color="B3452C")
ws.column_dimensions["B"].width = 110

# ============ Допущения ============
ws = wb.create_sheet("Допущения")
sheet_header(ws, "Допущения — единственный источник ввода", "Каждое число ниже имеет обоснование в колонке D; всё остальное в книге — формулы")
ws.column_dimensions["B"].width = 46; ws.column_dimensions["C"].width = 14; ws.column_dimensions["D"].width = 78
r = 4
r = section(ws, r, "СПРОС И ПОВЕДЕНИЕ КЛИЕНТА")
r = put(ws, r, "Средний чек (AOV), ₽", 1250, "aov", F_RUB, True, "Средний между «обед на одного» (~600) и «ужин на семью» (~2000); проверяется пилотом")
r = put(ws, r, "Заказов в месяц на активного клиента", 2.4, "orders_pm", '0.0', True, "Еда — частотная категория: 2–3 заказа/мес у нишевых фуд-сервисов; консервативно ниже агрегаторов (4–6)")
r = put(ws, r, "Удержание: доля месяца 2 (M2)", 0.86, "ret_decay", F_PCT, True, "Геометрический спад: каждый месяц остаётся 86% от предыдущего — до пола")
r = put(ws, r, "Удержание: пол (лояльное ядро)", 0.20, "ret_floor", F_PCT, True, "20% становятся постоянными («свой повар») — эффект локальной привычки")
r = put(ws, r, "Горизонт LTV, месяцев", 24, "horizon", F_NUM, True, "Стандартный горизонт для маркетплейсов на seed-стадии")
r += 1
r = section(ws, r, "ВЫРУЧКА ПЛАТФОРМЫ (по потокам)")
r = put(ws, r, "Комиссия с заказа", 0.10, "commission", F_PCT, True, "Зафиксирована в продукте (PLATFORM_FEE=0.1) — козырь против 25–35% агрегаторов")
r = put(ws, r, "SaaS-подписка повара, ₽/мес", 9900, "saas", F_RUB, True, "Тариф «Профи»: аналитика, продвижение, записи эфиров")
r = put(ws, r, "Заказов на повара в месяц", 160, "orders_per_chef", F_NUM, True, "5–6 заказов/день на активного повара — уровень занятости «частичный день»")
r = put(ws, r, "Featured-размещения, ₽/мес на повара", 2500, "featured", F_RUB, True, "Средний расход повара на буст карточки/эфира")
r = put(ws, r, "Премиум-подписка клиента, ₽/мес", 399, "premium_price", F_RUB, True, "Бесплатная доставка + ранний доступ к эфирам")
r = put(ws, r, "Доля клиентов на премиуме", 0.08, "premium_share", F_PCT, True, "Консервативно: у зрелых сервисов 10–15%")
r += 1
r = section(ws, r, "ПЕРЕМЕННЫЕ ЗАТРАТЫ НА ЗАКАЗ")
r = put(ws, r, "Эквайринг, % от чека", 0.026, "acquiring", F_PCT, True, "Ставка интернет-эквайринга ЮKassa/CloudPayments для малого бизнеса")
r = put(ws, r, "Поддержка и операции, ₽/заказ", 12, "support_per_order", F_RUB, True, "Амортизация саппорта: 1 обращение на ~25 заказов × себестоимость обращения ~300 ₽")
r = put(ws, r, "Доля менеджера района от комиссии", 0.10, "mgr_share", F_PCT, True, "Мотивация менеджера-партнёра — платится из комиссии платформы")
r += 1
r = section(ws, r, "CAC ПО КАНАЛАМ (веса и параметры)")
r = put(ws, r, "Таргет (VK/Директ): CPC, ₽", 28, "cpc", F_RUB, True, "Средний CPC фуд-тематики в городах-миллионниках")
r = put(ws, r, "Таргет: конверсия клик → регистрация", 0.07, "cr_click_reg", F_PCT, True, "Посадочная с живым эфиром; бенчмарк лендингов еды 5–10%")
r = put(ws, r, "Таргет: конверсия регистрация → 1-й заказ", 0.38, "cr_reg_order", F_PCT, True, "Приветственные 500 FC снижают порог первого заказа")
r = put(ws, r, "Рефералка: бонус обоим, ₽ (2×300 FC)", 600, "ref_bonus", F_RUB, True, "Начисляется после первого заказа приглашённого — не тратится вхолостую")
r = put(ws, r, "Рефералка: доля дошедших до заказа", 0.85, "ref_activation", F_PCT, True, "Бонус привязан к заказу, поэтому активация высокая")
r = put(ws, r, "Посевы в районе: бюджет, ₽/мес", 40000, "seed_budget", F_RUB, True, "Районные чаты, листовки с QR, дегустации у поваров")
r = put(ws, r, "Посевы: новых клиентов в месяц", 55, "seed_clients", F_NUM, True, "Оценка по охвату районных сообществ (5–15 тыс. человек)")
r = put(ws, r, "SEO-контент: бюджет, ₽/мес", 18000, "seo_budget", F_RUB, True, "Амортизация производства рецептов-посадочных (делает Claude)")
r = put(ws, r, "SEO: новых клиентов в месяц", 120, "seo_clients", F_NUM, True, "После индексации 100+ рецептов; маржинальная стоимость → 0")
r = put(ws, r, "Веса каналов: таргет / рефералка / посевы / SEO", "30/25/25/20", None, '@', True, "Структура привлечения на этапе пилота")
r = put(ws, r, "Вес: таргет", 0.30, "w_ads", F_PCT, True, "")
r = put(ws, r, "Вес: рефералка", 0.25, "w_ref", F_PCT, True, "")
r = put(ws, r, "Вес: посевы", 0.25, "w_seed", F_PCT, True, "")
r = put(ws, r, "Вес: SEO", 0.20, "w_seo", F_PCT, True, "")
r = put(ws, r, "Накладные на привлечение (аналитика, инструменты)", 0.12, "cac_overhead", F_PCT, True, "Надбавка к смешанному CAC")
r += 1
r = section(ws, r, "ЭКОНОМИКА РАЙОНА")
r = put(ws, r, "Целевые заказы/день района (плато)", 40, "district_target", F_NUM, True, "10 поваров × 4–6 заказов/день = целевая загрузка района")
r = put(ws, r, "Месяцев до плато", 6, "district_ramp", F_NUM, True, "S-образный разгон: посевы + сарафан + эфиры")
r = put(ws, r, "Маркетинг района, ₽/мес", 55000, "district_mkt", F_RUB, True, "Посевы 40к + доля таргета 15к")

# ============ Выручка на заказ ============
ws = wb.create_sheet("Выручка на заказ")
sheet_header(ws, "Выручка и вклад на один заказ", "Take-rate здесь — результат, а не допущение")
ws.column_dimensions["B"].width = 46; ws.column_dimensions["C"].width = 15; ws.column_dimensions["D"].width = 66
r = 4
r = section(ws, r, "ВЫРУЧКА ПЛАТФОРМЫ НА ЗАКАЗ")
r0 = r
r = put(ws, r, "Комиссия с заказа, ₽", f"={REF['aov']}*{REF['commission']}", None, F_RUB2, note="AOV × комиссия")
r = put(ws, r, "SaaS на заказ, ₽", f"={REF['saas']}/{REF['orders_per_chef']}", None, F_RUB2, note="подписка повара / его заказы в месяц")
r = put(ws, r, "Featured на заказ, ₽", f"={REF['featured']}/{REF['orders_per_chef']}", None, F_RUB2, note="буст-бюджет повара / его заказы")
r = put(ws, r, "Премиум на заказ, ₽", f"={REF['premium_share']}*{REF['premium_price']}/{REF['orders_pm']}", None, F_RUB2, note="доля премиум-клиентов × цена / заказы клиента в месяц")
rev_row = r
r = put(ws, r, "ИТОГО выручка платформы на заказ, ₽", f"=SUM(C{r0}:C{r-1})", None, F_RUB2)
ws.cell(rev_row, 3).font = RES_FONT
take_row = r
r = put(ws, r, "Эффективный take-rate (вычислен)", f"=C{rev_row}/{REF['aov']}", None, F_PCT, note="ранее в модели была экспертная константа 18,55% — теперь считается")
ws.cell(take_row, 3).font = RES_FONT
r += 1
r = section(ws, r, "ПЕРЕМЕННЫЕ ЗАТРАТЫ НА ЗАКАЗ")
c0 = r
r = put(ws, r, "Эквайринг, ₽", f"={REF['aov']}*{REF['acquiring']}", None, F_RUB2)
r = put(ws, r, "Поддержка и операции, ₽", f"={REF['support_per_order']}", None, F_RUB2)
r = put(ws, r, "Доля менеджера района, ₽", f"={REF['aov']}*{REF['commission']}*{REF['mgr_share']}", None, F_RUB2)
cost_row = r
r = put(ws, r, "ИТОГО затраты на заказ, ₽", f"=SUM(C{c0}:C{r-1})", None, F_RUB2)
r += 1
contrib_row = r
r = put(ws, r, "ВКЛАД (contribution) на заказ, ₽", f"=C{rev_row}-C{cost_row}", None, F_RUB2)
ws.cell(contrib_row, 3).font = RES_FONT
margin_row = r
r = put(ws, r, "Маржа платформы от своей выручки", f"=C{contrib_row}/C{rev_row}", None, F_PCT)
REF["rev_per_order"] = f"'Выручка на заказ'!$C${rev_row}"
REF["contrib_per_order"] = f"'Выручка на заказ'!$C${contrib_row}"
REF["take_rate"] = f"'Выручка на заказ'!$C${take_row}"

# ============ CAC ============
ws = wb.create_sheet("CAC")
sheet_header(ws, "Стоимость привлечения клиента — по каналам", "Смешанный CAC = взвешенная сумма каналов + накладные")
ws.column_dimensions["B"].width = 46; ws.column_dimensions["C"].width = 15; ws.column_dimensions["D"].width = 66
r = 4
r = section(ws, r, "СТОИМОСТЬ КАНАЛОВ")
ads_row = r
r = put(ws, r, "Таргет: CAC, ₽", f"={REF['cpc']}/({REF['cr_click_reg']}*{REF['cr_reg_order']})", None, F_RUB, note="CPC / (CR клик→рег × CR рег→заказ)")
ref_row = r
r = put(ws, r, "Рефералка: CAC, ₽", f"={REF['ref_bonus']}/{REF['ref_activation']}", None, F_RUB, note="бонусы / доля активации")
seed_row = r
r = put(ws, r, "Посевы района: CAC, ₽", f"={REF['seed_budget']}/{REF['seed_clients']}", None, F_RUB, note="бюджет / приведённые клиенты")
seo_row = r
r = put(ws, r, "SEO-контент: CAC, ₽", f"={REF['seo_budget']}/{REF['seo_clients']}", None, F_RUB, note="бюджет контента / клиенты из поиска")
r += 1
r = section(ws, r, "СМЕШАННЫЙ CAC")
blend_row = r
r = put(ws, r, "Взвешенный CAC, ₽", f"=C{ads_row}*{REF['w_ads']}+C{ref_row}*{REF['w_ref']}+C{seed_row}*{REF['w_seed']}+C{seo_row}*{REF['w_seo']}", None, F_RUB)
cac_row = r
r = put(ws, r, "CAC с накладными, ₽", f"=C{blend_row}*(1+{REF['cac_overhead']})", None, F_RUB)
ws.cell(cac_row, 3).font = RES_FONT
REF["cac"] = f"'CAC'!$C${cac_row}"

# ============ Когорты LTV ============
ws = wb.create_sheet("Когорты LTV")
sheet_header(ws, "Когорта из 100 клиентов на горизонте 24 месяцев", "Retention: M1=100%, далее ×decay до пола; LTV = накопленный вклад на клиента")
heads = ["Месяц", "Retention", "Активных из 100", "Заказов", "Вклад когорты, ₽", "LTV накопл., ₽/клиент"]
for j, h in enumerate(heads, start=2):
    c = ws.cell(4, j, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.border = BORDER
    ws.column_dimensions[get_column_letter(j)].width = 18
for m in range(1, 25):
    row = 4 + m
    ws.cell(row, 2, m).number_format = F_NUM
    if m == 1:
        ws.cell(row, 3, 1.0)
    else:
        ws.cell(row, 3, f"=MAX({REF['ret_floor']},C{row-1}*{REF['ret_decay']})")
    ws.cell(row, 3).number_format = F_PCT
    ws.cell(row, 4, f"=100*C{row}").number_format = '0.0'
    ws.cell(row, 5, f"=D{row}*{REF['orders_pm']}").number_format = '0.0'
    ws.cell(row, 6, f"=E{row}*{REF['contrib_per_order']}").number_format = F_RUB
    prev = f"G{row-1}+" if m > 1 else ""
    ws.cell(row, 7, f"={prev}F{row}/100").number_format = F_RUB
    for j in range(2, 8): ws.cell(row, j).border = BORDER
ltv_row = 4 + 24
ws.cell(ltv_row + 2, 2, "LTV на горизонте 24 мес, ₽/клиент").font = RES_FONT
ws.cell(ltv_row + 2, 3, f"=G{ltv_row}").number_format = F_RUB
ws.cell(ltv_row + 2, 3).font = RES_FONT
REF["ltv"] = f"'Когорты LTV'!$C${ltv_row + 2}"
ws.cell(ltv_row + 3, 2, "Месячный вклад активного клиента, ₽").font = Font(size=9, color="6D5E45")
ws.cell(ltv_row + 3, 3, f"={REF['orders_pm']}*{REF['contrib_per_order']}").number_format = F_RUB
REF["monthly_contrib"] = f"'Когорты LTV'!$C${ltv_row + 3}"

# ============ Юнит-экономика (сводка) ============
ws = wb.create_sheet("Юнит-экономика")
sheet_header(ws, "Сводка юнит-экономики", "Все значения — формулы; изменение любого допущения пересчитывает всё")
ws.column_dimensions["B"].width = 46; ws.column_dimensions["C"].width = 16; ws.column_dimensions["D"].width = 66
r = 4
r = section(ws, r, "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ")
r = put(ws, r, "AOV, ₽", f"={REF['aov']}", None, F_RUB)
r = put(ws, r, "Выручка платформы на заказ, ₽", f"={REF['rev_per_order']}", None, F_RUB2)
r = put(ws, r, "Эффективный take-rate", f"={REF['take_rate']}", None, F_PCT)
r = put(ws, r, "Вклад на заказ, ₽", f"={REF['contrib_per_order']}", None, F_RUB2)
r = put(ws, r, "LTV (24 мес), ₽", f"={REF['ltv']}", None, F_RUB)
r = put(ws, r, "CAC (смешанный), ₽", f"={REF['cac']}", None, F_RUB)
ltv_cac_row = r
r = put(ws, r, "LTV / CAC", f"={REF['ltv']}/{REF['cac']}", None, F_X)
ws.cell(ltv_cac_row, 3).font = RES_FONT
pb_row = r
r = put(ws, r, "Окупаемость клиента (payback)", f"={REF['cac']}/{REF['monthly_contrib']}", None, F_MO)
ws.cell(pb_row, 3).font = RES_FONT
r += 1
ws.cell(r, 2, "Критерии здоровья: LTV/CAC > 3× — хорошо; payback < 6 мес — хорошо.").font = Font(size=9, color="6D5E45")

# ============ Экономика района ============
ws = wb.create_sheet("Экономика района")
sheet_header(ws, "Район как юнит масштабирования", "S-образный разгон до плато; когда район выходит в плюс")
heads = ["Месяц", "Заказов/день", "Заказов/мес", "Вклад, ₽/мес", "Маркетинг, ₽/мес", "Прибыль района, ₽", "Накопленно, ₽"]
for j, h in enumerate(heads, start=2):
    c = ws.cell(4, j, h); c.fill = HDR_FILL; c.font = HDR_FONT; c.border = BORDER
    ws.column_dimensions[get_column_letter(j)].width = 17
for m in range(1, 13):
    row = 4 + m
    ws.cell(row, 2, m).number_format = F_NUM
    # S-кривая: target * m^2/(m^2 + (ramp/2)^2)
    ws.cell(row, 3, f"={REF['district_target']}*B{row}^2/(B{row}^2+({REF['district_ramp']}/2)^2)").number_format = '0.0'
    ws.cell(row, 4, f"=C{row}*30").number_format = F_NUM
    ws.cell(row, 5, f"=D{row}*{REF['contrib_per_order']}").number_format = F_RUB
    ws.cell(row, 6, f"={REF['district_mkt']}").number_format = F_RUB
    ws.cell(row, 7, f"=E{row}-F{row}").number_format = F_RUB
    prev = f"H{row-1}+" if m > 1 else ""
    ws.cell(row, 8, f"={prev}G{row}").number_format = F_RUB
    for j in range(2, 9): ws.cell(row, j).border = BORDER
ws.cell(18, 2, "Прибыль района на плато, ₽/мес").font = RES_FONT
ws.cell(18, 3, "=G16").number_format = F_RUB; ws.cell(18, 3).font = RES_FONT
ws.cell(19, 2, "Месяц выхода в операционный плюс").font = Font(size=9, color="6D5E45")
ws.cell(19, 3, '=MATCH(TRUE,INDEX(G5:G16>0,0),0)').number_format = F_NUM

# ============ Чувствительность ============
ws = wb.create_sheet("Чувствительность")
sheet_header(ws, "LTV/CAC при разных удержании и чеке", "Сетка формул: строки — retention decay, столбцы — AOV")
decays = [0.80, 0.83, 0.86, 0.89, 0.92]
aovs = [1000, 1250, 1500]
# помощник: сумма retention по decay (24 мес, пол из Допущений)
ws.cell(4, 2, "Помощник: Σ retention (24 мес)").font = SEC_FONT
for i, d in enumerate(decays):
    row = 5 + i
    ws.cell(row, 2, d).number_format = F_PCT
    # первая ячейка месяца
    ws.cell(row, 3, 1.0)
    for m in range(2, 25):
        col = 2 + m
        ws.cell(row, col, f"=MAX({REF['ret_floor']},{get_column_letter(col-1)}{row}*$B${row})")
    ws.cell(row, 28, f"=SUM(C{row}:AA{row})").number_format = '0.00'
ws.cell(4, 28, "Σ").font = SEC_FONT
# матрица LTV/CAC: contribution пересчитывается на AOV (комиссия и эквайринг зависят от чека)
base = 12
ws.cell(base, 2, "LTV / CAC").fill = HDR_FILL; ws.cell(base, 2).font = HDR_FONT
for j, a in enumerate(aovs):
    c = ws.cell(base, 3 + j, a); c.number_format = F_RUB; c.fill = HDR_FILL; c.font = HDR_FONT
for i, d in enumerate(decays):
    row = base + 1 + i
    ws.cell(row, 2, d).number_format = F_PCT
    for j, a in enumerate(aovs):
        col = 3 + j
        aov_cell = f"{get_column_letter(3 + j)}{base}"
        sum_cell = f"$AB${5 + i}"
        contrib = (f"({aov_cell}*{REF['commission']}"
                   f"+{REF['saas']}/{REF['orders_per_chef']}"
                   f"+{REF['featured']}/{REF['orders_per_chef']}"
                   f"+{REF['premium_share']}*{REF['premium_price']}/{REF['orders_pm']}"
                   f"-{aov_cell}*{REF['acquiring']}"
                   f"-{REF['support_per_order']}"
                   f"-{aov_cell}*{REF['commission']}*{REF['mgr_share']})")
        ws.cell(row, col, f"={REF['orders_pm']}*{contrib}*{sum_cell}/{REF['cac']}").number_format = F_X
        ws.cell(row, col).border = BORDER
ws.cell(base + 7, 2, "Зелёная зона > 3×; базовый сценарий — центр сетки (decay 86%, AOV 1250).").font = Font(size=9, color="6D5E45")
for cletter in ["B"]: ws.column_dimensions[cletter].width = 16
for j in range(3, 6): ws.column_dimensions[get_column_letter(j)].width = 12

wb.save("ForkWork-Unit-Economics-v2.xlsx")
print("OK: ForkWork-Unit-Economics-v2.xlsx")
